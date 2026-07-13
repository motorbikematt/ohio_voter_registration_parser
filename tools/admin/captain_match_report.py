"""captain_match_report.py -- P0 match report for Democratic precinct captains.

The first deliverable of the Quorum seeder build (HANDOFF_SEEDER_BUILD_2026-06-30
sections 3-4). It answers ONE operational question before any registry is seeded:

    Of the Democratic precinct captains, how many resolve to an exact voter
    record, how many only fuzzy-match, and how many do NOT match at all?

The operator sets the confidence bar AFTER seeing this distribution -- this script
sets no bar; it MEASURES (handoff section 3).

SINGLE-RESOLVER RULE (CLAUDE.md section 5). This report does NOT re-implement any
matching logic. It imports and calls the one voter-matcher that already exists in
match_to_voters.py:

    * load_voter_index(slug)                 -> county voter index (bucketed by surname)
    * match_entity(index, last, first, ...)  -> the ONLY fuzzy matcher (its two-pass
                                                exact-prefix + rapidfuzz strategy,
                                                the `location_check` states)
    * derive_match_confidence(...)           -> the high / medium / low tiers
    * match_key(entity_type, identity)       -> ledger join key (section 2a: this is
                                                the same construct as electedofficials_key)
    * binding_hash(row)                      -> point-in-time voter fingerprint (2d)

Captains come from the AUTHORITATIVE officials file (handoff section 1):
    local/source/County Data Files/57_Montgomery/electedofficials.csv
    rows where DISTTYPE=PRECINCT and PARTY=D (176 rows in the current drop).

PRECINCT IDENTITY comes from DISTNAME + the precinct-keys crosswalk, NEVER from
parsing OFCDESC (the OFCDESC bug -- SESSION_FINDINGS_2026-06-29 section 9). The
crosswalked PRECINCT_NAME is the location constraint handed to match_entity, exactly
as the officials pipeline's own captain path does.

The five corroborating fields the handoff cites (Name + Precinct + Address + Zip +
City) are NOT all in electedofficials.csv -- its office-address block (OADDR1/OCITY/
OSTATE/OZIPCODE) is uniformly BLANK for captains. Name + Precinct drive the match;
the residential Zip lives only in the secondary cross-check roster
(Other/Copy of Dem CC.xlsx). We attach that zip best-effort as a SECONDARY
corroboration of the matched voter -- it never drives identity or the match itself.

LOUD + GRACEFUL (CLAUDE.md section 7; handoff section 3). An unmatched captain is
logged to the report with a valid electedofficials_key and "SOS_VOTERID pending";
it is never silently dropped and never halts the run.

Usage:
    python tools/admin/captain_match_report.py
    python tools/admin/captain_match_report.py --county montgomery --verbose
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import sys
from pathlib import Path

# tools/admin/ on sys.path so the sibling resolver + crosswalk import cleanly.
sys.path.insert(0, str(Path(__file__).resolve().parent))

from officials_common import (  # noqa: E402
    COUNTY_NUMBER,
    ROOT,
    atomic_write_json,
    load_precinct_crosswalk,
    name_from_parts,
)
from match_to_voters import (  # noqa: E402  -- REUSED resolver, not re-implemented
    binding_hash,
    derive_match_confidence,
    load_ledger,
    load_voter_index,
    match_entity,
    match_key,
    verification_projection,
)

COUNTY_DIR = ROOT / "local" / "source" / "County Data Files" / "57_Montgomery"
ELECTED_CSV = COUNTY_DIR / "electedofficials.csv"
DEM_CC_XLSX = COUNTY_DIR / "Other" / "Copy of Dem CC.xlsx"
OUT_DIR = ROOT / "local" / "working"

ENTITY_TYPE = "captain_candidate"  # match the officials pipeline's captain bucket
OFCDESC_PREFIX = "DEMOCRATIC CENTRAL COMMITTEE "


# -- captains from the authoritative CSV ---------------------------------------

def load_dem_captains() -> list[dict]:
    """DEM precinct-captain rows from electedofficials.csv.

    IDENTIFY DEMOCRATIC CAPTAINS BY THE OFCDESC PARTY PREFIX, NOT BY PARTY.
    The PARTY column is unreliable for committee seats -- the same behavior-derived /
    can-be-blank rule that governs SWVF PARTY_AFFILIATION (CLAUDE.md section 4) shows
    up here: of the 209 'DEMOCRATIC CENTRAL COMMITTEE' precinct rows, 175 are coded
    D, 33 are BLANK, and 1 is miscoded R -- a PARTY='D' filter silently dropped 34
    real Democratic captains. So the population filter is:
        upper(trim(OFCDESC)) startswith 'DEMOCRATIC CENTRAL COMMITTEE'
        AND upper(trim(DISTTYPE)) == 'PRECINCT'   -> exactly 209.

    This does NOT violate the 'never parse OFCDESC for identity' rule (that rule is
    about the precinct-name SUFFIX, which had the OFCDESC bug). Using the party
    PREFIX ('DEMOCRATIC' vs 'REPUBLICAN CENTRAL COMMITTEE') to determine party is
    clean and reliable. Precinct IDENTITY still comes from DISTNAME + the crosswalk.

    Blank SWVF/CSV cells are empty strings, not null (CLAUDE.md section 4) -- every
    field is read defensively as (value or "").strip().
    """
    prefix = OFCDESC_PREFIX.strip()
    rows = list(csv.DictReader(ELECTED_CSV.open(encoding="utf-8")))
    dem: list[dict] = []
    for r in rows:
        if (r.get("DISTTYPE") or "").strip().upper() != "PRECINCT":
            continue
        if not (r.get("OFCDESC") or "").strip().upper().startswith(prefix):
            continue
        dem.append(r)
    return dem


def electedofficials_key(row: dict) -> str:
    """The always-present captain identity spine (handoff section 2a).

    "<DISTNAME>|<PARTY>|<LASTN>,<FIRSTN>,<MIDDLEN>", built from verbatim CSV fields
    so a human can grep the source and land on the row. Raw DISTNAME (the reliable
    precinct code), never OFCDESC. Same construct as match_to_voters.match_key.
    """
    dn = (row.get("DISTNAME") or "").strip()
    party = (row.get("PARTY") or "").strip()
    last = (row.get("LASTN") or "").strip()
    first = (row.get("FIRSTN") or "").strip()
    middle = (row.get("MIDDLEN") or "").strip()
    return f"{dn}|{party}|{last},{first},{middle}"


def _ofcdesc_abbrev(row: dict) -> str | None:
    """The precinct abbreviation suffix in OFCDESC (e.g. 'BRK-A').

    Used ONLY to join the secondary DEM CC roster's zip onto a captain -- NEVER to
    derive precinct identity (that is DISTNAME + crosswalk). Returns None when the
    suffix is absent (the OFCDESC bug), which just drops the optional zip enrichment.
    """
    d = (row.get("OFCDESC") or "").strip().upper()
    if d.startswith(OFCDESC_PREFIX):
        suffix = d[len(OFCDESC_PREFIX):].strip()
        return suffix or None
    return None


# -- secondary residential-zip cross-check (Copy of Dem CC.xlsx) ---------------

def load_dem_cc_zip_by_abbrev() -> dict[str, dict]:
    """{precinct_abbrev -> {zip, city, address, name}} from the DEM CC roster.

    Best-effort secondary corroboration only. If the file or openpyxl is missing we
    return {} and the report proceeds without zip corroboration (graceful, section 5).
    """
    if not DEM_CC_XLSX.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        print("NOTE: openpyxl not installed; skipping DEM CC zip cross-check.",
              file=sys.stderr)
        return {}

    wb = openpyxl.load_workbook(DEM_CC_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out: dict[str, dict] = {}
    # header: Precinct/Ward/Township | Name | Address | Zip Code | City
    for r in rows[1:]:
        if not r or not (r[1] and str(r[1]).strip()):
            continue
        abbrev = (str(r[0]).strip().upper() if r[0] is not None else "")
        z = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
        # Excel stores the zip as an int; re-pad to 5.
        z = z.split(".")[0].zfill(5) if z else ""
        out[abbrev] = {
            "name": str(r[1]).strip() if r[1] else "",
            "address": str(r[2]).strip() if len(r) > 2 and r[2] else "",
            "zip": z,
            "city": str(r[4]).strip() if len(r) > 4 and r[4] else "",
        }
    return out


# -- report --------------------------------------------------------------------

def run(county_slug: str, verbose: bool) -> dict:
    try:
        from rapidfuzz import fuzz, process
    except ImportError:
        print("ERROR: rapidfuzz not installed. Run: uv add rapidfuzz", file=sys.stderr)
        sys.exit(1)

    index = load_voter_index(county_slug)              # REUSED resolver machinery
    n_voters = sum(len(v) for v in index.values())
    print(f"Loaded {n_voters:,} ACTIVE/CONFIRMATION voters for {county_slug} "
          f"({len(index):,} distinct surnames)")

    crosswalk = load_precinct_crosswalk(county_slug, validate=True)
    print(f"Precinct crosswalk: {len(crosswalk.ballot_to_name):,} ballot numbers")

    captains = load_dem_captains()
    print(f"DEM precinct captains (OFCDESC 'DEMOCRATIC CENTRAL COMMITTEE%', "
          f"DISTTYPE=PRECINCT): {len(captains)}")

    cc_zip = load_dem_cc_zip_by_abbrev()
    if cc_zip:
        print(f"DEM CC secondary zip roster: {len(cc_zip)} precincts")
    else:
        print("DEM CC secondary zip roster: unavailable (zip cross-check skipped)")

    ledger = load_ledger()                                    # C7: durable confirmations
    by_id = {r["SOS_VOTERID"]: r for rows_ in index.values() for r in rows_}

    tiers = {"high": 0, "medium": 0, "low": 0}
    methods: dict[str, int] = {}
    # Party-coding audit: how each captain was coded in the raw PARTY column, given
    # they were ALL selected by the OFCDESC 'DEMOCRATIC CENTRAL COMMITTEE' prefix.
    # "D" = coded D; "" (blank) = the 33 formerly-dropped captains; "R" = miscoded.
    party_coding: dict[str, int] = {}
    matched: list[dict] = []
    unmatched: list[dict] = []
    unresolved_precinct: list[dict] = []
    zip_agree = zip_disagree = zip_no_signal = 0
    verification_summary = {"unverified": 0, "confirmed": 0, "rejected": 0, "corrected": 0}

    for row in captains:
        eo_key = electedofficials_key(row)
        distname = (row.get("DISTNAME") or "").strip()
        last = (row.get("LASTN") or "").strip()
        first = (row.get("FIRSTN") or "").strip()
        middle = (row.get("MIDDLEN") or "").strip()
        suffix = (row.get("SUFFIXN") or "").strip()
        raw_party = (row.get("PARTY") or "").strip()   # verbatim; may be "", "D", "R"
        party_coding[raw_party or "(blank)"] = party_coding.get(raw_party or "(blank)", 0) + 1
        display_name = name_from_parts(first, middle, last, suffix)

        # Precinct identity strictly from DISTNAME + crosswalk (never OFCDESC).
        precinct_name = crosswalk.resolve(distname)
        record = {
            "electedofficials_key": eo_key,
            "distname": distname,
            "name": display_name,
            "last_name": last,
            "first_name": first,
            "precinct_name": precinct_name,
            # raw PARTY cell; captain is Democratic by OFCDESC prefix regardless.
            "raw_party": raw_party,
            "party_miscoded": raw_party not in ("", "D"),
        }
        # match_key computed for EVERY captain (moved out of the matched-only
        # branch) so a ledger `corrected` entry can promote an otherwise
        # unmatched captain into the matched list, not just adjust one already
        # matched (same fix as seed_quorum_registry.py Part 2).
        mk = match_key(ENTITY_TYPE, record)
        record["match_key"] = mk
        entry = ledger.get(mk)
        verification_summary[entry.get("state", "unverified") if entry else "unverified"] += 1

        if precinct_name is None:
            # Crosswalk miss: loud, but the captain keeps a valid identity and we
            # still attempt a name-only match (constraint=None).
            record["issue"] = "distname_not_in_precinct_crosswalk"
            unresolved_precinct.append(record)
            constraint = None
        else:
            constraint = ("precinct", precinct_name)

        # THE reused matcher -- no fuzzy logic is written here.
        vrow, method, score, location_check = match_entity(
            index, last, first, constraint, fuzz, process)

        # C7: a human `corrected` entry overrides the auto-match with the right
        # voter -- same override pattern as match_to_voters.py's _record(), so
        # this can promote a captain the matcher itself couldn't find.
        if entry and entry.get("state") == "corrected":
            cid = entry.get("corrected_sos_voterid")
            if cid in by_id:
                vrow, method, score, location_check = by_id[cid], "human_corrected", 100, "confirmed"
            elif verbose:
                print(f"  NOTE  ledger correction {cid!r} not found in voter index "
                      f"for {display_name}")

        methods[method] = methods.get(method, 0) + 1

        if vrow is None:
            record["match_method"] = method
            record["match_confidence"] = None
            record["sos_voterid"] = None
            record["sos_voterid_status"] = "pending"   # loud + graceful (section 3)
            record["verification"] = verification_projection(
                entry, {"sos_voterid": None, "binding_hash": None})
            unmatched.append(record)
            if verbose:
                print(f"  MISS  {display_name:28s} [{precinct_name}] ({method})")
            continue

        confidence = derive_match_confidence(method, location_check, None)
        tiers[confidence] += 1

        # Secondary zip corroboration (never gates the match).
        abbrev = _ofcdesc_abbrev(row)
        cc = cc_zip.get(abbrev) if abbrev else None
        cc_z = (cc or {}).get("zip") or ""
        voter_zip = (vrow.get("RESIDENTIAL_ZIP") or "").strip()
        if cc_z and voter_zip:
            if cc_z[:5] == voter_zip[:5]:
                zip_check = "agree"
                zip_agree += 1
            else:
                zip_check = "disagree"
                zip_disagree += 1
        else:
            zip_check = "no_signal"
            zip_no_signal += 1

        vbhash = binding_hash(vrow)
        record.update({
            "match_method": method,
            "match_score": score,
            "location_check": location_check,
            "match_confidence": confidence,
            "sos_voterid": vrow.get("SOS_VOTERID"),
            "sos_voterid_status": "matched",
            "voter_last_name": vrow.get("LAST_NAME"),
            "voter_first_name": vrow.get("FIRST_NAME"),
            "voter_precinct_name": vrow.get("PRECINCT_NAME"),
            "voter_residential_zip": voter_zip or None,
            "dem_cc_zip": cc_z or None,
            "zip_cross_check": zip_check,
            "binding_hash": vbhash,
            "verification": verification_projection(
                entry, {"sos_voterid": vrow.get("SOS_VOTERID"), "binding_hash": vbhash}),
        })
        matched.append(record)
        if verbose:
            print(f"  HIT   {display_name:28s} [{method}/{score}] "
                  f"-> {record['sos_voterid']} ({confidence}, zip:{zip_check})")

    missing_sos = unmatched  # every unmatched captain lacks a SOS_VOTERID
    meta = {
        "generated": _dt.datetime.now().isoformat(timespec="seconds"),
        "county": county_slug,
        "county_number": COUNTY_NUMBER[county_slug],
        "entity_type": ENTITY_TYPE,
        "sources": {
            "captains_authoritative": str(ELECTED_CSV.relative_to(ROOT)),
            "precinct_crosswalk": f"local/source/precinct_keys/{county_slug}_precincts.csv",
            "voters": "local/source/parquet_enriched/enriched_voters.parquet",
            "dem_cc_secondary_zip": str(DEM_CC_XLSX.relative_to(ROOT)) if cc_zip else None,
        },
        "resolver_reused": {
            "module": "tools/admin/match_to_voters.py",
            "functions": ["load_voter_index", "match_entity",
                          "derive_match_confidence", "match_key", "binding_hash",
                          "load_ledger", "verification_projection"],
            "note": "no matching logic re-implemented; single-resolver rule (CLAUDE.md 5)",
        },
        "verification_summary": verification_summary,
        "totals": {
            "dem_captains": len(captains),
            "matched": len(matched),
            "unmatched": len(unmatched),
            "missing_sos_voterid": len(missing_sos),
            "precinct_unresolved": len(unresolved_precinct),
        },
        "by_confidence": tiers,
        "by_method": methods,
        # Raw PARTY coding of the OFCDESC-selected Democratic captains. The 33
        # "(blank)" + any miscoded value are the captains a PARTY='D' filter would
        # have silently dropped (coordinator correction, 2026-07-01).
        "party_coding": party_coding,
        "population_filter": "OFCDESC startswith 'DEMOCRATIC CENTRAL COMMITTEE' AND DISTTYPE=PRECINCT",
        "zip_cross_check": {
            "agree": zip_agree, "disagree": zip_disagree, "no_signal": zip_no_signal,
        },
    }
    return {
        "_meta": meta,
        "missing_sos_voterid": [
            {"electedofficials_key": r["electedofficials_key"],
             "name": r["name"], "distname": r["distname"],
             "precinct_name": r["precinct_name"], "match_method": r["match_method"],
             "raw_party": r["raw_party"]}
            for r in missing_sos
        ],
        "matched": matched,
        "unmatched": unmatched,
        "precinct_unresolved": unresolved_precinct,
    }


def write_text_report(result: dict, path: Path) -> None:
    """Human-readable companion (ASCII only -- Windows cp1252 console safe)."""
    m = result["_meta"]
    t = m["totals"]
    lines: list[str] = []
    lines.append("DEM PRECINCT CAPTAIN -> VOTER MATCH REPORT")
    lines.append("=" * 52)
    lines.append(f"generated : {m['generated']}")
    lines.append(f"county    : {m['county']} ({m['county_number']})")
    lines.append(f"captains  : {t['dem_captains']} "
                 f"(OFCDESC 'DEMOCRATIC CENTRAL COMMITTEE%', DISTTYPE=PRECINCT)")
    lines.append("")
    lines.append("RAW PARTY CODING (all selected by OFCDESC prefix)")
    for code, n in sorted(m["party_coding"].items(),
                          key=lambda kv: (kv[0] != "D", kv[0])):
        note = ""
        if code == "(blank)":
            note = "  <- formerly dropped by PARTY='D' filter"
        elif code not in ("D",):
            note = "  <- miscoded (Democratic by OFCDESC), included"
        lines.append(f"  PARTY={code:8s}: {n}{note}")
    lines.append("")
    lines.append("MATCH DISTRIBUTION")
    lines.append(f"  matched            : {t['matched']}")
    lines.append(f"  unmatched          : {t['unmatched']}")
    lines.append(f"  missing SOS_VOTERID: {t['missing_sos_voterid']}")
    lines.append("")
    lines.append("CONFIDENCE TIERS (over matched)")
    for tier in ("high", "medium", "low"):
        lines.append(f"  {tier:6s}: {m['by_confidence'][tier]}")
    lines.append("")
    lines.append("MATCH METHOD")
    for meth, n in sorted(m["by_method"].items(), key=lambda kv: -kv[1]):
        lines.append(f"  {meth:20s}: {n}")
    lines.append("")
    z = m["zip_cross_check"]
    lines.append("SECONDARY ZIP CROSS-CHECK (DEM CC roster vs matched voter)")
    lines.append(f"  agree    : {z['agree']}")
    lines.append(f"  disagree : {z['disagree']}")
    lines.append(f"  no_signal: {z['no_signal']}")
    lines.append("")
    if m["totals"]["precinct_unresolved"]:
        lines.append(f"PRECINCT CROSSWALK MISSES: {m['totals']['precinct_unresolved']}")
        for r in result["precinct_unresolved"]:
            lines.append(f"  {r['distname']:6s} {r['name']}")
        lines.append("")
    lines.append(f"CAPTAINS MISSING SOS_VOTERID ({len(result['missing_sos_voterid'])})")
    if not result["missing_sos_voterid"]:
        lines.append("  (none)")
    for r in result["missing_sos_voterid"]:
        pc = r.get("raw_party") or "(blank)"
        lines.append(f"  {r['distname']:6s} {r['name']:30s} "
                     f"precinct={r['precinct_name']}  ({r['match_method']}, PARTY={pc})")
    lines.append("")
    vs = m.get("verification_summary", {})
    lines.append(f"LEDGER VERIFICATION ({sum(vs.values())} captains checked)")
    for state in ("unverified", "confirmed", "rejected", "corrected"):
        lines.append(f"  {state:10s}: {vs.get(state, 0)}")
    lines.append("")
    ledger_flagged = [
        r for r in (result["matched"] + result["unmatched"])
        if r.get("verification", {}).get("status") in ("stale", "rejected")
        or (r.get("verification", {}).get("status") == "unverified"
            and r.get("match_confidence") in (None, "low"))
    ]
    lines.append(f"LEDGER-FLAGGED NEEDS REVIEW ({len(ledger_flagged)})")
    if not ledger_flagged:
        lines.append("  (none)")
    for r in sorted(ledger_flagged, key=lambda r: r["distname"]):
        vstatus = r.get("verification", {}).get("status", "unverified")
        conf = r.get("match_confidence") or "unmatched"
        lines.append(f"  {r['distname']:6s} {r['name']:30s} "
                     f"verification={vstatus}  confidence={conf}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="P0 match report: DEM precinct captains -> voter records.")
    ap.add_argument("--county", default="montgomery")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    if args.county not in COUNTY_NUMBER:
        print(f"[abort] unknown county slug {args.county!r}; known: {list(COUNTY_NUMBER)}",
              file=sys.stderr)
        return 1

    result = run(args.county, args.verbose)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = OUT_DIR / f"captain_match_report_{args.county}.json"
    txt_path = OUT_DIR / f"captain_match_report_{args.county}.txt"
    nbytes = atomic_write_json(json_path, result)
    write_text_report(result, txt_path)

    m = result["_meta"]
    t = m["totals"]
    bc = m["by_confidence"]
    pc = m["party_coding"]
    print("\n" + "=" * 52)
    print(f"DEM captains          : {t['dem_captains']}  "
          f"(raw PARTY: " + ", ".join(f"{k}={v}" for k, v in sorted(pc.items())) + ")")
    print(f"  matched             : {t['matched']}  "
          f"[high {bc['high']} / medium {bc['medium']} / low {bc['low']}]")
    print(f"  unmatched           : {t['unmatched']}")
    print(f"  missing SOS_VOTERID  : {t['missing_sos_voterid']}")
    z = m["zip_cross_check"]
    print(f"  zip cross-check      : agree {z['agree']} / "
          f"disagree {z['disagree']} / no_signal {z['no_signal']}")
    print(f"\nWrote {json_path.relative_to(ROOT)} ({nbytes:,} bytes)")
    print(f"Wrote {txt_path.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
