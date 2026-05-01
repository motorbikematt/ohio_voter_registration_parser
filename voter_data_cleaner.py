import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
INPUT_CSV = r"D:\vibe\election-data\source\2026-04-30\voterfile.csv"
OUTPUT_XLSX = r"D:\vibe\election-data\voter_analysis.xlsx"

def clean_voter_data(df):
    """Clean and prepare voter data"""
    print("Cleaning voter data...")

    # Birth year validation
    df['BIRTHYEAR'] = pd.to_numeric(df['BIRTHYEAR'], errors='coerce')

    # Remove impossible birth years (before 1900, after current year)
    df = df[(df['BIRTHYEAR'] >= 1900) & (df['BIRTHYEAR'] <= 2024)]

    # Create decade grouping column
    df['Decade'] = (df['BIRTHYEAR'] // 10 * 10).astype(int)

    # Clean address fields
    df['full_address'] = (
        df['STNUM'].astype(str) + " " +
        df['STDIR'].fillna('') + " " +
        df['STNAME'].astype(str) + " " +
        df['APT'].fillna('')
    ).str.strip().str.replace(r'\s+', ' ', regex=True)

    # Standardize party affiliation
    df['PARTYAFFIL'] = df['PARTYAFFIL'].fillna('Unknown').str.strip()

    return df

def create_excel_workbook(df, output_path):
    """Create XLSX with cleaned data and summary table"""
    print(f"Creating workbook: {output_path}")

    wb = Workbook()

    # Sheet 1: Cleaned Voter Data
    ws_data = wb.active
    ws_data.title = "Voter Data"

    # Write headers
    headers = df.columns.tolist()
    ws_data.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width

    # Sheet 2: Decade Summary for Histogram
    ws_summary = wb.create_sheet("Decade Summary")

    # Get unique decades and sort
    decades = sorted(df['Decade'].dropna().unique())

    # Write summary headers
    ws_summary['A1'] = "Decade"
    ws_summary['B1'] = "Voter Count"

    summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for cell in ['A1', 'B1']:
        ws_summary[cell].fill = summary_header_fill
        ws_summary[cell].font = summary_header_font
        ws_summary[cell].alignment = Alignment(horizontal="center", vertical="center")

    # Write decade labels and formulas for counting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, decade in enumerate(decades, 2):
        ws_summary[f'A{idx}'] = decade
        ws_summary[f'A{idx}'].alignment = Alignment(horizontal="center")
        ws_summary[f'A{idx}'].border = thin_border

        # Use COUNTIFS formula to count voters in this decade from raw data
        ws_summary[f'B{idx}'] = f'=COUNTIFS(\'Voter Data\'!$BIRTHYEAR:$BIRTHYEAR,">="&{decade},\'Voter Data\'!$BIRTHYEAR:$BIRTHYEAR,"<"&{decade+10})'
        ws_summary[f'B{idx}'].alignment = Alignment(horizontal="right")
        ws_summary[f'B{idx}'].number_format = '#,##0'
        ws_summary[f'B{idx}'].border = thin_border

    # Format summary columns
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 15

    # Save workbook
    wb.save(output_path)
    print(f"✓ Workbook saved to: {output_path}")
    print(f"✓ Sheet 1 'Voter Data': All {len(df)} cleaned voter records")
    print(f"✓ Sheet 2 'Decade Summary': {len(decades)} decade cohorts with formulas ready for charting")

def main():
    print("=" * 60)
    print("VOTER DATA CLEANING & XLSX EXPORT")
    print("=" * 60)

    # Check input file
    if not os.path.exists(INPUT_CSV):
        print(f"✗ Error: Input file not found: {INPUT_CSV}")
        return

    # Load data
    print(f"Loading: {INPUT_CSV}")
    try:
        df = pd.read_csv(INPUT_CSV)
        print(f"✓ Loaded {len(df):,} rows, {len(df.columns)} columns")
    except Exception as e:
        print(f"✗ Error reading CSV: {e}")
        return

    # Clean data
    try:
        df = clean_voter_data(df)
        print(f"✓ Cleaned: {len(df):,} valid voter records")
    except Exception as e:
        print(f"✗ Error cleaning data: {e}")
        return

    # Create Excel workbook
    try:
        create_excel_workbook(df, OUTPUT_XLSX)
        print("\n" + "=" * 60)
        print("SUCCESS! Ready to create histogram in Excel")
        print("=" * 60)
    except Exception as e:
        print(f"✗ Error creating workbook: {e}")
        return

if __name__ == "__main__":
    main()
